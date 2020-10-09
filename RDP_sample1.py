from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Color, Alignment, Border, Side, colors, NamedStyle, PatternFill
import pandas as pd
import refinitiv.dataplatform as rdp
import logging.config

# set your RDP credentials here. 
# See RDP Quickstart for description - https://developers.refinitiv.com/en/api-catalog/refinitiv-data-platform/refinitiv-data-platform-apis/quick-start
RDP_LOGIN                   = '--- YOUR RDP USER ID ---'
RDP_PASSWORD                = '--- YOUR RDP PASSWORD ---'
APP_KEY                     = '--- YOUR GENERATED APP KEY ---'

INPUT_PORTFOLIO				= 'InputPortfolio.xlsx'
OUTPUT_PORTFOLIO			= 'output.xlsx'


print('Loading input portfolio...')
workbook = load_workbook(filename=INPUT_PORTFOLIO)

sheet = workbook.active
data = sheet.values
cols = next(data)
data = list(data)

inputdf = pd.DataFrame(data, columns=cols)
inputdf.dropna(inplace=True)
print(inputdf.head())

# Create an RDP session and get data
print('Logging into RDP...')
session = rdp.open_platform_session(
	APP_KEY, 
	rdp.GrantPassword(
		username = RDP_LOGIN, 
		password = RDP_PASSWORD
	)
)

print('Getting ESG data for portfolio...')

endpoint = rdp.Endpoint(
	session = rdp.get_default_session(),
	url = "data/datagrid/beta1/")

instrs = inputdf['Instrument'].values.tolist()

response = endpoint.send_request(
	method = rdp.Endpoint.RequestMethod.POST,
	body_parameters = {
		"fields": [
			"TR.TRESGScore",
			"TR.TRBCEconomicSector",
			"TR.ExchangeCountry",
			"TR.ExchangeRegion"
		],
		"universe": instrs
	}
)

if response.is_success:
	headers = [h['name'] for h in response.data.raw['headers']]
	dataf = pd.DataFrame(data=response.data.raw['data'], columns=headers)
	print(dataf.head())
else:
	print('Error: Unable to get the data for the portfolio')
	print(response.error_message)
	exit()
	
rdp.close_session()

# data retrieved, post process
dataf.rename(columns = {'instrument':'Instrument'}, inplace = True)
dataf.rename(columns = {'TR.ExchangeCountry':'Exchange Country'}, inplace = True) 
dataf.rename(columns = {'TR.TRBCEconomicSector':'TRBC Economic Sector'}, inplace = True) 
dataf.rename(columns = {'TR.TRESGScore':'ESG Score'}, inplace = True) 
dataf.rename(columns = {'TR.ExchangeRegion':'Exchange Region'}, inplace = True)

df = pd.merge(inputdf, dataf)

#% of portfolio with ESG coverage = nn%
df_esg_sum = df[df['ESG Score'].notna()].copy()
df_esg_sum = df_esg_sum['Portfolio Weight'].sum()
print('Number of holdings with ESG coverage: {:.2%}'.format(df_esg_sum))

#Holdings without ESG
df_esg_na = df[df['ESG Score'].isnull()].copy()
df_esg_na['Portfolio Weight'] = pd.Series(["{0:.3f}%".format(val * 100) for val in df_esg_na['Portfolio Weight']], index = df_esg_na.index)
del df_esg_na[None]
df_esg_na

#Top & bottom 5 holdings by weight (which have ESG coverage)
df_top = df.nlargest(5,'Portfolio Weight')
df_top = df_top[['Instrument','Issuer Name','Portfolio Weight','ESG Score','TRBC Economic Sector']]
df_top['Portfolio Weight'] = pd.Series(["{0:.3f}%".format(val * 100) for val in df_top['Portfolio Weight']], index = df_top.index)
df_bottom = df.nsmallest(5,'Portfolio Weight')
df_bottom = df_bottom[['Instrument','Issuer Name','Portfolio Weight','ESG Score','TRBC Economic Sector']]
df_bottom['Portfolio Weight'] = pd.Series(["{0:.3f}%".format(val * 100) for val in df_bottom['Portfolio Weight']], index = df_bottom.index)

#Top & bottom 5 holdings by ESG score
df_esg_port = df[df['ESG Score'].notna()].copy()
df_esg_port['ESG Portfolio Weight'] = df_esg_port['Portfolio Weight'].transform(lambda x: x / x.sum())

df_esg_top = df_esg_port.nlargest(5,'ESG Score')
df_esg_top = df_esg_top[['Instrument','Issuer Name','Portfolio Weight','ESG Score','TRBC Economic Sector']]
df_esg_bottom = df_esg_port.nsmallest(5,'ESG Score')
df_esg_bottom = df_esg_bottom[['Instrument','Issuer Name','Portfolio Weight','ESG Score','TRBC Economic Sector']]

def wavg(group, avg_name, weight_name):
	d = group[avg_name]
	w = group[weight_name]
	try:
		return (d * w).sum() / w.sum()
	except ZeroDivisionError:
		return d.mean()

df_esg_region = df_esg_port.groupby(['Exchange Region']).apply(wavg, 'ESG Score','ESG Portfolio Weight')
df_esg_region = df_esg_region.to_frame('ESG Score').reset_index()
df_esg_region['ESG Score'] = df_esg_region['ESG Score'].apply(lambda x: round(x, 3))
df_region_allo = df_esg_port.groupby(['Exchange Region'])['ESG Portfolio Weight'].sum()
df_region_allo = df_region_allo.to_frame('Portfolio Weight').reset_index()
df_region_allo['Portfolio Weight'] = pd.Series(["{0:.3f}%".format(val * 100) for val in df_region_allo['Portfolio Weight']], index = df_region_allo.index)

# ESG score by country
df_esg_country = df_esg_port.groupby(['Exchange Country']).apply(wavg, 'ESG Score','ESG Portfolio Weight')
df_esg_country = df_esg_country.to_frame('ESG Score').reset_index()
df_esg_country['ESG Score'] = df_esg_country['ESG Score'].apply(lambda x: round(x, 3))
# Portfolio Allocation by Country
df_country_allo = df_esg_port.groupby(['Exchange Country'])['ESG Portfolio Weight'].sum()
df_country_allo = df_country_allo.to_frame('Portfolio Weight').reset_index()
df_country_allo['Portfolio Weight'] = pd.Series(["{0:.3f}%".format(val * 100) for val in df_country_allo['Portfolio Weight']], index = df_country_allo.index)

# ESG sunburst chart

print('Creating a new Excel spreadsheet...')

# Write into a new Excel spreadsheet
def resizeColumns(sht):
	dims = {}
	for row in sht.rows:
		for cell in row:
			if cell.value:
				dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))   
	for col, value in dims.items():
		sht.column_dimensions[col].width = value

bold_font = Font(bold=True)
bkgClr = PatternFill(fgColor="E0E0E0", fill_type = "solid")
#define header style
header = NamedStyle(name="header")
header.font = bold_font
header.border = Border(bottom=Side(border_style="thin"))
header.alignment = Alignment(horizontal="center", vertical="center")

def addDataFrame(theSht, tableTitle, sourceDF):
	curRow = theSht._current_row
	curCell = 'A' + str(curRow + 1)
	theSht[curCell] = tableTitle
	theSht[curCell].font = bold_font
	theSht[curCell].fill = bkgClr
	for row in dataframe_to_rows(sourceDF, index=False, header=True):
		theSht.append(row)
	header_row = theSht[curRow + 2]
	for cell in header_row:
		cell.style = header

workbook = Workbook()

# First sheet with raw data
ws1 = workbook.active
ws1.title = 'Portfolio'

addDataFrame(ws1, 'Complete Portfolio', df)
resizeColumns(ws1)

# Holdings without ESG coverage
ws2 = workbook.create_sheet("No ESG Coverage")
ws2['A1'] = 'Number of holdings without ESG coverage: {:.2%}'.format(1 - df_esg_sum)
ws2['A1'].font = bold_font
ws2.append([])
ws2.append([])
addDataFrame(ws2, 'Holdings without any ESG coverage', df_esg_na)
resizeColumns(ws2)

#Top & bottom 5 holdings by weight
ws3 = workbook.create_sheet("By Weight")
addDataFrame(ws3, 'Top 5 holdings by weight', df_top)
ws3.append([])
ws3.append([])
addDataFrame(ws3, 'Bottom 5 holdings by weight', df_bottom)
resizeColumns(ws3)

#Top & bottom 5 holdings by ESG score
ws4 = workbook.create_sheet("By ESG Score")
addDataFrame(ws4, 'Top 5 holdings by ESC Score', df_esg_top)
ws4.append([])
ws4.append([])
addDataFrame(ws4, 'Bottom 5 holdings by ESG Score', df_esg_bottom)
resizeColumns(ws4)

#Portfolio Allocation and ESG Score by Region
ws5 = workbook.create_sheet("Geography")
addDataFrame(ws5, 'Portfolio Allocation by Region', df_region_allo)
ws5.append([])
ws5.append([])
addDataFrame(ws5, 'Portfolio Allocation by Country', df_country_allo)
ws5.append([])
ws5.append([])
addDataFrame(ws5, 'ESG Score by Region', df_esg_region)
ws5.append([])
ws5.append([])
addDataFrame(ws5, 'ESG Score by Country', df_esg_country)
resizeColumns(ws5)

workbook.save(OUTPUT_PORTFOLIO)

print('Finished!')
